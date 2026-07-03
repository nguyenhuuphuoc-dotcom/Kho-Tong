"""
routers/import_data.py — Import hàng loạt từ file Excel (sheet QLTK)
"""
from fastapi import APIRouter, HTTPException, UploadFile, File, Form
from typing import Optional
import openpyxl, io, re, time
import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
import supabase_client as db

router = APIRouter(prefix="/api/import", tags=["import"])


def _to_date(val, fallback="2025-01-01"):
    if val is None:
        return fallback
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    return s[:10] if len(s) >= 10 else fallback


def _parse_qltk(ws):
    """Parse sheet QLTK, trả về (hang_hoa_list, nhap_groups, xuat_groups)."""

    # Danh mục (cột A:B)
    hang_hoa_list = []
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if row[0] and str(row[0]).strip():
            hang_hoa_list.append({
                "ten": str(row[0]).strip(),
                "dvt": str(row[1]).strip() if row[1] else "cái",
            })

    # Nhập kho (cột H:L)
    nhap_groups = {}
    last_date = None
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not row[7]:
            continue
        if row[9]:
            last_date = row[9]
        ngay = _to_date(last_date)
        ncc = str(row[11]).strip() if row[11] else ""
        key = (ngay, ncc)
        nhap_groups.setdefault(key, []).append({
            "ten_hang": str(row[7]).strip(),
            "dvt": str(row[8]).strip() if row[8] else "cái",
            "so_luong": float(row[10]) if row[10] else 0,
            "don_gia": 0, "thanh_tien": 0, "ghi_chu": "",
        })

    # Xuất kho (cột N:R)
    xuat_groups = {}
    last_date = None
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
        if not row[13]:
            continue
        if row[15]:
            last_date = row[15]
        ngay = _to_date(last_date)
        nguoi = str(row[17]).strip() if row[17] else ""
        key = (ngay, nguoi)
        xuat_groups.setdefault(key, []).append({
            "ten_hang": str(row[13]).strip(),
            "dvt": str(row[14]).strip() if row[14] else "cái",
            "so_luong": float(row[16]) if row[16] else 0,
            "don_gia": 0, "thanh_tien": 0, "ghi_chu": "",
        })

    return hang_hoa_list, nhap_groups, xuat_groups


@router.post("/preview")
async def preview_import(
    file: UploadFile = File(...),
    cong_trinh_id: int = Form(...),
):
    """Đọc file Excel, trả về thống kê — KHÔNG insert dữ liệu."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không đọc được file Excel: {str(e)}")

    if "QLTK" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="File không có sheet QLTK")

    ws = wb["QLTK"]
    hang_hoa_list, nhap_groups, xuat_groups = _parse_qltk(ws)

    return {
        "ok": True,
        "hang_hoa": len(hang_hoa_list),
        "phieu_nk": len(nhap_groups),
        "dong_nk": sum(len(v) for v in nhap_groups.values()),
        "phieu_xk": len(xuat_groups),
        "dong_xk": sum(len(v) for v in xuat_groups.values()),
    }


@router.post("/execute")
async def execute_import(
    file: UploadFile = File(...),
    cong_trinh_id: int = Form(...),
):
    """Import thật sự: insert hang_hoa + phieu NK/XK vào Supabase."""
    try:
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Không đọc được file Excel: {str(e)}")

    if "QLTK" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="File không có sheet QLTK")

    ws = wb["QLTK"]
    hang_hoa_list, nhap_groups, xuat_groups = _parse_qltk(ws)

    hh_ok = hh_err = nk_ok = nk_err = xk_ok = xk_err = 0

    # ── 1. Danh mục hàng hóa ─────────────────────────────────
    existing = set()
    try:
        rows = db.select("hang_hoa",
                         query="ten_hang",
                         filters=f"cong_trinh_id=eq.{cong_trinh_id}")
        existing = {r["ten_hang"] for r in rows}
    except Exception:
        pass

    BATCH = 50
    to_insert = [
        {
            "ma_hang": f"HH-{i+1:04d}-{int(time.time()) % 10000}",
            "ten_hang": h["ten"],
            "dvt": h["dvt"],
            "nhom": "Vật tư",
            "cong_trinh_id": cong_trinh_id,
        }
        for i, h in enumerate(hang_hoa_list)
        if h["ten"] not in existing
    ]

    for i in range(0, len(to_insert), BATCH):
        batch = to_insert[i:i + BATCH]
        try:
            db.insert("hang_hoa", batch)
            hh_ok += len(batch)
        except Exception:
            for item in batch:
                try:
                    item["ma_hang"] += f"-{hh_err}"
                    db.insert("hang_hoa", item)
                    hh_ok += 1
                except Exception:
                    hh_err += 1

    # ── 2. Phiếu Nhập kho ────────────────────────────────────
    for idx, ((ngay, ncc), items) in enumerate(nhap_groups.items()):
        so_phieu = f"NK-IMP-{idx+1:04d}"
        try:
            phieu = db.create_phieu(
                cong_trinh_id=cong_trinh_id,
                loai="NK",
                so_phieu=so_phieu,
                ngay=ngay,
                doi_tac=ncc,
                ghi_chu="Import từ Excel",
                tong_tien=0,
                nguon="import",
            )
            if phieu:
                db.push_chi_tiet(phieu["id"], items)
                nk_ok += 1
            else:
                nk_err += 1
        except Exception:
            nk_err += 1

    # ── 3. Phiếu Xuất kho ────────────────────────────────────
    for idx, ((ngay, nguoi), items) in enumerate(xuat_groups.items()):
        so_phieu = f"XK-IMP-{idx+1:04d}"
        try:
            phieu = db.create_phieu(
                cong_trinh_id=cong_trinh_id,
                loai="XK",
                so_phieu=so_phieu,
                ngay=ngay,
                doi_tac=nguoi,
                ghi_chu="Import từ Excel",
                tong_tien=0,
                nguon="import",
            )
            if phieu:
                db.push_chi_tiet(phieu["id"], items)
                xk_ok += 1
            else:
                xk_err += 1
        except Exception:
            xk_err += 1

    return {
        "ok": True,
        "hang_hoa": {"thanh_cong": hh_ok, "loi": hh_err},
        "nhap_kho":  {"thanh_cong": nk_ok, "loi": nk_err},
        "xuat_kho":  {"thanh_cong": xk_ok, "loi": xk_err},
    }
